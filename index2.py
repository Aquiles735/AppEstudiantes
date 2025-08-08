import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Importa las clases de las otras ventanas
from componentes.notas import VentanaNotas
from componentes.modifica import VentanaModificar

from componentes.diseno import Diseno 
from componentes.widgets import Widgets 

class Control:
    db_name = 'registro_estudiante.db'

    def __init__(self, window):
        self.wind = window
        self.wind.title('Sistema de Registro de Estudiantes (Elab: Prof.Aquiles M. 08/2025)')
        self.wind.geometry('1000x700')
        self.wind.config(bg='#2c3e50') 

        self.diseno = Diseno(self.wind)
        self.ui = Widgets(self.wind, self)
        
        
        # Carga los estudiantes en el Treeview al iniciar
        self.get_estudiantes()

    def run_query(self, query, parameters=()):
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            result = cursor.execute(query, parameters)
            conn.commit()
        return result

    def get_estudiantes(self):
        records = self.ui.tree.get_children()
        for element in records:
            self.ui.tree.delete(element)

        query = 'SELECT cedula_estudiante, nombre, apellido, nivel, seccion FROM estudiantes ORDER BY nivel ASC, seccion ASC, CAST(cedula_estudiante AS INTEGER) ASC;'
        db_rows = self.run_query(query)
        for row in db_rows:
            self.ui.tree.insert('', 'end', text=row[0], values=row[0:])

    # LÓGICA DE VALIDACIÓN
    def validation(self):
        self.ui.messaje['text'] = ''
        cedula = self.ui.cedula_entry.get()
        nombre = self.ui.nombre_entry.get()
        apellido = self.ui.apellido_entry.get()

        if not cedula.isdigit():
            self.ui.messaje['text'] = 'La Cédula debe contener solo números.'
            return False

        if not nombre or not apellido:
            self.ui.messaje['text'] = 'Todos los campos son requeridos.'
            return False

        return True

    # REGISTRAR ESTUDIANTE
    import sqlite3
    def resgist_estud(self):
        if self.validation():
            cedula_ingresada = self.ui.cedula_entry.get()
            
            #Revision si la cedula existe
            check_query = 'SELECT cedula_estudiante FROM estudiantes WHERE cedula_estudiante = ?'
            existing_student = self.run_query(check_query, (cedula_ingresada,)).fetchone()

            # Si el estudiante ya existe, muestra el error y limpia todos los campos
            if existing_student:
                self.ui.messaje.config(fg='#e74c3c')
                self.ui.messaje['text'] = 'Error: Un estudiante con esa cédula ya existe.'
                self.ui.cedula_entry.delete(0, 'end')
                self.ui.nombre_entry.delete(0, 'end')
                self.ui.apellido_entry.delete(0, 'end')
                self.ui.nivel_combobox.set('')
                self.ui.seccion_combobox.set('')
                
            else:
                # Si no existe, procede con la inserción y limpia todos los campos
                query = 'INSERT INTO estudiantes VALUES (?, ?, ?, ?, ?)'
                parameters = (
                    cedula_ingresada,
                    self.ui.nombre_entry.get(),
                    self.ui.apellido_entry.get(),
                    self.ui.nivel_combobox.get(),
                    self.ui.seccion_combobox.get()
                )
                try:
                    self.run_query(query, parameters)
                    self.ui.messaje.config(fg='#2ecc71')
                    self.ui.messaje['text'] = f'Estudiante {self.ui.nombre_entry.get()} agregado satisfactoriamente'
                    self.get_estudiantes()

                    # Limpia los campos después de una inserción exitosa
                    self.ui.cedula_entry.delete(0, 'end')
                    self.ui.nombre_entry.delete(0, 'end')
                    self.ui.apellido_entry.delete(0, 'end')
                   

                except Exception as e:
                    self.ui.messaje.config(fg='#e74c3c')
                    self.ui.messaje['text'] = f'Ocurrió un error inesperado: {e}'
                
        else:
            self.ui.messaje.config(fg='#e74c3c')
        


    # CARGAR DATOS EN CAMPOS DE ENTRADA
    def seleccionar_estudiante(self, event):
        self.ui.messaje['text'] = ''
        selected_item = self.ui.tree.focus()
        if not selected_item:
            return

        item_data = self.ui.tree.item(selected_item)
        values = item_data['values']

        self.ui.cedula_entry.config(state='normal')
        self.ui.cedula_entry.delete(0, tk.END)
        self.ui.cedula_entry.insert(0, values[0])
        self.ui.cedula_entry.config(state='readonly')
        
        self.ui.nombre_entry.delete(0, tk.END)
        self.ui.nombre_entry.insert(0, values[1])
        
        self.ui.apellido_entry.delete(0, tk.END)
        self.ui.apellido_entry.insert(0, values[2])
        
        self.ui.nivel_combobox.set(values[3])
        
        self.ui.seccion_combobox.set(values[4])

    # AGREGAR NOTAS
    def agregar_notas(self):
        self.ui.messaje['text'] = ''
        try:
            selected_item = self.ui.tree.focus()
            if not selected_item:
                messagebox.showwarning("Error", "Debe seleccionar un estudiante para agregar notas.")
                return

            data = self.ui.tree.item(selected_item)
            if 'values' in data:
                cedula_estudiante = data['values'][0]
                nivel_estudiante = data['values'][3]
                seccion_estudiante = data['values'][4]
                
                VentanaNotas(
                    self.wind,
                    cedula_estudiante=cedula_estudiante,
                    nivel_estudiante=nivel_estudiante,
                    seccion_estudiante=seccion_estudiante,
                    callback_actualizar=self.get_estudiantes
                )
            else:
                self.ui.messaje.config(fg='#e74c3c')
                self.ui.messaje['text'] = 'Seleccione un estudiante de la tabla para registrar notas.'
        except IndexError:
            self.ui.messaje.config(fg='#e74c3c')
            self.ui.messaje['text'] = 'Seleccione un estudiante de la tabla para registrar notas.'

    # MODIFICAR ESTUDIANTE
    def modificar_estudiante(self):
        self.ui.messaje['text'] = ''
        try:
            selected_item = self.ui.tree.focus()
            values = self.ui.tree.item(selected_item, 'values')

            if not selected_item:
                self.ui.messaje.config(fg='#e74c3c')
                self.ui.messaje['text'] = 'Seleccione un estudiante de la tabla para modificar.'
                return

            if values:
                cedula_estudiante = values[0]
                nombre = values[1]
                apellido = values[2]
                nivel = values[3]
                seccion = values[4]
                
                VentanaModificar(
                    self.wind, 
                    cedula_estudiante_orig=cedula_estudiante,
                    nombre_orig=nombre,
                    apellido_orig=apellido,
                    nivel_orig=nivel,
                    seccion_orig=seccion,
                    callback_actualizar=self.get_estudiantes
                )
        
        except IndexError:
            self.ui.messaje.config(fg='#e74c3c')
            self.ui.messaje['text'] = 'Seleccione un estudiante de la tabla para modificar.'
        except Exception as e:
            self.ui.messaje.config(fg='#e74c3c')
            self.ui.messaje['text'] = f'Ocurrió un error inesperado: {e}'

    # ELIMINAR ESTUDIANTE
    def eliminar_estudiante(self):
        self.ui.messaje['text'] = ''
        try:
            selected_item = self.ui.tree.selection()
            if not selected_item:
                messagebox.showwarning("Error", "Debe seleccionar un estudiante para eliminar.")
                return

            cedula_a_eliminar = self.ui.tree.item(selected_item, 'text')

            respuesta = messagebox.askyesno(
                "Confirmar Eliminación",
                f"¿Está seguro de que desea eliminar al estudiante con cédula {cedula_a_eliminar}?"
            )

            if respuesta:
                query = 'DELETE FROM estudiantes WHERE cedula_estudiante = ?'
                self.run_query(query, (cedula_a_eliminar,))

                query_notas = 'DELETE FROM notas WHERE cedula_estudiante = ?'
                self.run_query(query_notas, (cedula_a_eliminar,))

                self.ui.messaje.config(fg='#2ecc71')
                self.ui.messaje['text'] = f'Estudiante con cédula {cedula_a_eliminar} eliminado correctamente.'
                self.get_estudiantes()
        except Exception as e:
            self.ui.messaje.config(fg='#e74c3c')
            self.ui.messaje['text'] = f'Error al eliminar el estudiante: {e}'

    # DESCARGAR NOTAS A EXCEL
    def descargar_notas_a_excel(self):
        try:
            selected_item = self.ui.tree.focus()
            if not selected_item:
                messagebox.showwarning("Error", "Debe seleccionar un estudiante para descargar notas de la sección.")
                return

            values = self.ui.tree.item(selected_item, 'values')
            if not values:
                messagebox.showwarning("Error", "Los datos del estudiante seleccionado no son válidos.")
                return

            nivel_seleccionado = values[3]
            seccion_seleccionada = values[4]
            
            query = """
            SELECT
                e.cedula_estudiante, e.nombre, e.apellido, e.nivel, e.seccion,
                n.evaluacion_1, n.evaluacion_2, n.evaluacion_3, n.evaluacion_4, n.evaluacion_5,
                n.evaluacion_6, n.evaluacion_7, n.evaluacion_8, n.evaluacion_9, n.evaluacion_10,
                n.promedio_notas, n.nota_definitiva
            FROM estudiantes e
            INNER JOIN notas n ON e.cedula_estudiante = n.cedula_estudiante
            WHERE e.nivel = ? AND e.seccion = ?
            ORDER BY e.nivel ASC, e.seccion ASC, CAST(e.cedula_estudiante AS INTEGER) ASC;
            """
            parameters = (nivel_seleccionado, seccion_seleccionada)
            data_rows = self.run_query(query, parameters).fetchall()

            if not data_rows:
                messagebox.showinfo("Información", "No hay datos para descargar para el nivel y sección seleccionados.")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = f"Notas {nivel_seleccionado} {seccion_seleccionada}"

            headers = [
                'Cédula', 'Nombre', 'Apellido', 'Nivel', 'Sección',
                'Eval 1', 'Eval 2', 'Eval 3', 'Eval 4', 'Eval 5',
                'Eval 6', 'Eval 7', 'Eval 8', 'Eval 9', 'Eval 10',
                'Promedio', 'TOTAL'
            ]
            ws.append(headers)
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            row_number = 2
            for row in data_rows:
                ws.append(row)
                
                for col_index in range(6, 18):
                    cell_to_align = ws.cell(row=row_number, column=col_index)
                    cell_to_align.alignment = Alignment(horizontal='left', vertical='center')
                row_number += 1
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Archivos de Excel", "*.xlsx")],
                title="Guardar como",
                initialfile=f"Notas_{nivel_seleccionado}_{seccion_seleccionada}.xlsx"
            )

            if filename:
                wb.save(filename)
                messagebox.showinfo("Éxito", f"Notas guardadas exitosamente en:\n{os.path.basename(filename)}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al descargar el archivo: {e}")

    def borrar_todas_las_notas(self):
        respuesta = messagebox.askyesno(
            "Confirmar",
            "¿Está seguro de que desea borrar todas las notas de todos los estudiantes? Esta acción es irreversible."
        )
        if respuesta:
            query = 'DELETE FROM notas'
            self.run_query(query)
            messagebox.showinfo("Éxito", "Todas las notas han sido borradas correctamente.")
            self.get_estudiantes()

    def borrar_todo(self):
        respuesta = messagebox.askyesno(
            "Confirmar",
            "ADVERTENCIA: ¿Está seguro de que desea borrar TODOS los estudiantes y TODAS sus notas? Esta acción es irreversible."
        )
        if respuesta:
            query_estudiantes = 'DELETE FROM estudiantes'
            self.run_query(query_estudiantes)
            query_notas = 'DELETE FROM notas'
            self.run_query(query_notas)
            messagebox.showinfo("Éxito", "Todos los estudiantes y sus notas han sido borrados correctamente.")
            self.get_estudiantes()

if __name__ == '__main__':
    window = tk.Tk()
    application = Control(window)
    window.mainloop()
    # mis anotaciones de mergencia por cambio linux----window A.M.
   #    python3 index2.py
   #    python3 -m venv venv     linux
   #    source venv/bin/activate    linux
   #    deactivate
   #   git checkout main     <para cambiar de rama a main>  o
   #   git checkout master
   #   git push origin HEAD     para push en master github
   #                                git pull origin master
   #     pyinstaller index2.py   pyinstaller --onefile index2.py 
   #AppAcademi/dist/index2 AppAcademi/dist/registro_estudiante.db
#      pyinstaller --onefile --noconsole index2.py
  #   /home/aquiles/Documentos/MiAcademi/AppAcademi/venv/bin/python

